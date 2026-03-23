"""
쿠팡 광고 보고서 데이터 모델 및 엑셀 파서.
"""
import logging
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

logger = logging.getLogger(__name__)


# ── 데이터 모델 ──

@dataclass
class CoupangCampaignStats:
    """쿠팡 캠페인 성과."""
    campaign_name: str = ""
    campaign_type: str = ""          # AI스마트, 수동퍼포먼스, 매출최적화 등
    status: str = ""                 # 진행중, 일시중지, 종료
    impressions: int = 0
    clicks: int = 0
    ctr: float = 0.0
    cost: int = 0                    # 집행광고비 (원)
    conversions: int = 0             # 전환수
    conversion_sales: int = 0        # 전환매출 (원)
    roas: float = 0.0               # 광고수익률 (%)
    total_sales: int = 0            # 전체 판매매출 (원)


@dataclass
class CoupangProductStats:
    """쿠팡 상품별 성과."""
    product_name: str = ""
    product_id: str = ""
    campaign_name: str = ""
    impressions: int = 0
    clicks: int = 0
    ctr: float = 0.0
    cost: int = 0
    conversions: int = 0
    conversion_sales: int = 0
    roas: float = 0.0
    total_sales: int = 0


@dataclass
class CoupangKeywordStats:
    """쿠팡 키워드별 성과."""
    keyword: str = ""
    campaign_name: str = ""
    product_name: str = ""
    match_type: str = ""             # 정확매칭, 구문매칭 등
    impressions: int = 0
    clicks: int = 0
    ctr: float = 0.0
    cost: int = 0
    conversions: int = 0
    conversion_sales: int = 0
    roas: float = 0.0


@dataclass
class CoupangReport:
    """쿠팡 광고 종합 보고서."""
    account_name: str = ""
    report_date: date = None
    date_from: date = None
    date_to: date = None

    # 전체 요약
    total_impressions: int = 0
    total_clicks: int = 0
    total_ctr: float = 0.0
    total_cost: int = 0
    total_conversions: int = 0
    total_conversion_sales: int = 0
    total_roas: float = 0.0
    total_sales: int = 0

    # 상세 데이터
    campaigns: list[CoupangCampaignStats] = field(default_factory=list)
    products: list[CoupangProductStats] = field(default_factory=list)
    keywords: list[CoupangKeywordStats] = field(default_factory=list)

    # AI 인사이트
    insights: list[str] = field(default_factory=list)


# ── 엑셀 파서 ──

# 쿠팡 엑셀 컬럼 매핑 (한글 → 영문 필드)
COLUMN_MAP = {
    # 캠페인 관련
    "캠페인": "campaign_name",
    "캠페인명": "campaign_name",
    "캠페인 이름": "campaign_name",
    "캠페인유형": "campaign_type",
    "캠페인 유형": "campaign_type",
    "광고유형": "campaign_type",
    "상태": "status",

    # 상품 관련
    "상품명": "product_name",
    "상품 이름": "product_name",
    "상품ID": "product_id",
    "상품 ID": "product_id",

    # 키워드 관련
    "키워드": "keyword",
    "검색어": "keyword",
    "매칭유형": "match_type",
    "매칭 유형": "match_type",

    # 성과 지표
    "노출수": "impressions",
    "노출": "impressions",
    "광고노출수": "impressions",
    "클릭수": "clicks",
    "클릭": "clicks",
    "광고클릭수": "clicks",
    "CTR": "ctr",
    "클릭률": "ctr",
    "광고비": "cost",
    "집행광고비": "cost",
    "집행 광고비": "cost",
    "광고비(원)": "cost",
    "총비용": "cost",
    "총 비용": "cost",
    "전환수": "conversions",
    "전환": "conversions",
    "총 전환수": "conversions",
    "광고 전환 판매수": "conversions",
    "광고 전환 주문수": "orders",
    "전환매출": "conversion_sales",
    "전환매출액": "conversion_sales",
    "총 전환매출액(14일)": "conversion_sales",
    "광고전환매출": "conversion_sales",
    "광고 전환 매출": "conversion_sales",
    "전환매출(14일)": "conversion_sales",
    "중요 지표": "conversion_sales",
    "ROAS": "roas",
    "광고수익률": "roas",
    "광고수익률(%)": "roas",
    "광고비 효율성": "roas",
    "전체매출": "total_sales",
    "전체 판매매출": "total_sales",
    "판매매출": "total_sales",
    "광고 당 매출": "total_sales",
    "전환율": "ctr",
    "최근 예산 실적": "budget_status",
    "목표 광고수익률": "target_roas",
}


def _safe_int(val) -> int:
    """값을 안전하게 int로 변환."""
    if val is None:
        return 0
    try:
        if isinstance(val, str):
            val = (val.replace(",", "").replace("원", "").replace("%", "")
                   .replace("회", "").replace("평균", "").replace("개", "").strip())
            if not val or val == "-":
                return 0
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _safe_float(val) -> float:
    """값을 안전하게 float로 변환."""
    if val is None:
        return 0.0
    try:
        if isinstance(val, str):
            val = (val.replace(",", "").replace("%", "").replace("원", "")
                   .replace("회", "").replace("평균", "").replace("개", "").strip())
            if not val or val == "-":
                return 0.0
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def parse_coupang_excel(file_path: str) -> CoupangReport:
    """
    쿠팡 광고 엑셀 파일을 파싱하여 CoupangReport 생성.

    Args:
        file_path: 엑셀 파일 경로

    Returns:
        CoupangReport 객체
    """
    import openpyxl

    report = CoupangReport(report_date=date.today())
    path = Path(file_path)

    if not path.exists():
        logger.error(f"파일이 존재하지 않습니다: {file_path}")
        return report

    logger.info(f"쿠팡 엑셀 파싱: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # 헤더 찾기 (첫 번째 비어있지 않은 행)
            header_idx = 0
            for i, row in enumerate(rows):
                if row and any(cell is not None for cell in row):
                    header_idx = i
                    break

            headers = [str(h).strip() if h else "" for h in rows[header_idx]]
            logger.info(f"시트 '{sheet_name}' 헤더: {headers}")

            # 컬럼 매핑
            col_map = {}
            for col_idx, header in enumerate(headers):
                if header in COLUMN_MAP:
                    col_map[COLUMN_MAP[header]] = col_idx

            logger.info(f"매핑된 컬럼: {col_map}")

            # 데이터 파싱
            data_rows = rows[header_idx + 1:]

            # 시트 유형 자동 감지
            has_keyword = "keyword" in col_map
            has_product = "product_name" in col_map or "product_id" in col_map
            has_campaign = "campaign_name" in col_map

            if has_keyword:
                _parse_keyword_rows(report, data_rows, col_map)
            elif has_product:
                _parse_product_rows(report, data_rows, col_map)
            elif has_campaign:
                _parse_campaign_rows(report, data_rows, col_map)
            else:
                # 기본: 캠페인으로 처리
                _parse_campaign_rows(report, data_rows, col_map)

        wb.close()

    except Exception as e:
        logger.error(f"엑셀 파싱 오류: {e}", exc_info=True)

    # 전체 요약 계산
    _calculate_totals(report)

    # 인사이트 생성
    report.insights = generate_coupang_insights(report)

    logger.info(
        f"파싱 완료: 캠페인 {len(report.campaigns)}개, "
        f"상품 {len(report.products)}개, "
        f"키워드 {len(report.keywords)}개"
    )

    return report


def _parse_campaign_rows(report: CoupangReport, rows: list, col_map: dict):
    """캠페인 행 파싱."""
    for row in rows:
        if not row or all(cell is None for cell in row):
            continue
        stats = CoupangCampaignStats(
            campaign_name=str(row[col_map["campaign_name"]]).strip() if "campaign_name" in col_map and row[col_map["campaign_name"]] else "",
            campaign_type=str(row[col_map.get("campaign_type", -1)]).strip() if col_map.get("campaign_type", -1) >= 0 and len(row) > col_map.get("campaign_type", 0) and row[col_map.get("campaign_type", 0)] else "",
            status=str(row[col_map.get("status", -1)]).strip() if col_map.get("status", -1) >= 0 and len(row) > col_map.get("status", 0) and row[col_map.get("status", 0)] else "",
            impressions=_safe_int(row[col_map["impressions"]] if "impressions" in col_map else 0),
            clicks=_safe_int(row[col_map["clicks"]] if "clicks" in col_map else 0),
            ctr=_safe_float(row[col_map["ctr"]] if "ctr" in col_map else 0),
            cost=_safe_int(row[col_map["cost"]] if "cost" in col_map else 0),
            conversions=_safe_int(row[col_map["conversions"]] if "conversions" in col_map else 0),
            conversion_sales=_safe_int(row[col_map["conversion_sales"]] if "conversion_sales" in col_map else 0),
            roas=_safe_float(row[col_map["roas"]] if "roas" in col_map else 0),
            total_sales=_safe_int(row[col_map["total_sales"]] if "total_sales" in col_map else 0),
        )
        if stats.campaign_name:
            report.campaigns.append(stats)


def _parse_product_rows(report: CoupangReport, rows: list, col_map: dict):
    """상품 행 파싱."""
    for row in rows:
        if not row or all(cell is None for cell in row):
            continue
        stats = CoupangProductStats(
            product_name=str(row[col_map["product_name"]]).strip() if "product_name" in col_map and row[col_map["product_name"]] else "",
            product_id=str(row[col_map["product_id"]]).strip() if "product_id" in col_map and col_map["product_id"] < len(row) and row[col_map["product_id"]] else "",
            campaign_name=str(row[col_map["campaign_name"]]).strip() if "campaign_name" in col_map and col_map["campaign_name"] < len(row) and row[col_map["campaign_name"]] else "",
            impressions=_safe_int(row[col_map["impressions"]] if "impressions" in col_map else 0),
            clicks=_safe_int(row[col_map["clicks"]] if "clicks" in col_map else 0),
            ctr=_safe_float(row[col_map["ctr"]] if "ctr" in col_map else 0),
            cost=_safe_int(row[col_map["cost"]] if "cost" in col_map else 0),
            conversions=_safe_int(row[col_map["conversions"]] if "conversions" in col_map else 0),
            conversion_sales=_safe_int(row[col_map["conversion_sales"]] if "conversion_sales" in col_map else 0),
            roas=_safe_float(row[col_map["roas"]] if "roas" in col_map else 0),
            total_sales=_safe_int(row[col_map["total_sales"]] if "total_sales" in col_map else 0),
        )
        if stats.product_name or stats.product_id:
            report.products.append(stats)


def _parse_keyword_rows(report: CoupangReport, rows: list, col_map: dict):
    """키워드 행 파싱."""
    for row in rows:
        if not row or all(cell is None for cell in row):
            continue
        stats = CoupangKeywordStats(
            keyword=str(row[col_map["keyword"]]).strip() if "keyword" in col_map and row[col_map["keyword"]] else "",
            campaign_name=str(row[col_map["campaign_name"]]).strip() if "campaign_name" in col_map and col_map["campaign_name"] < len(row) and row[col_map["campaign_name"]] else "",
            product_name=str(row[col_map["product_name"]]).strip() if "product_name" in col_map and col_map["product_name"] < len(row) and row[col_map["product_name"]] else "",
            match_type=str(row[col_map["match_type"]]).strip() if "match_type" in col_map and col_map["match_type"] < len(row) and row[col_map["match_type"]] else "",
            impressions=_safe_int(row[col_map["impressions"]] if "impressions" in col_map else 0),
            clicks=_safe_int(row[col_map["clicks"]] if "clicks" in col_map else 0),
            ctr=_safe_float(row[col_map["ctr"]] if "ctr" in col_map else 0),
            cost=_safe_int(row[col_map["cost"]] if "cost" in col_map else 0),
            conversions=_safe_int(row[col_map["conversions"]] if "conversions" in col_map else 0),
            conversion_sales=_safe_int(row[col_map["conversion_sales"]] if "conversion_sales" in col_map else 0),
            roas=_safe_float(row[col_map["roas"]] if "roas" in col_map else 0),
        )
        if stats.keyword:
            report.keywords.append(stats)


def _calculate_totals(report: CoupangReport):
    """전체 요약 계산."""
    # 캠페인 합산 (가장 상위 단위)
    source = report.campaigns or report.products or []
    for item in source:
        report.total_impressions += item.impressions
        report.total_clicks += item.clicks
        report.total_cost += item.cost
        report.total_conversions += item.conversions
        report.total_conversion_sales += item.conversion_sales
        if hasattr(item, "total_sales"):
            report.total_sales += item.total_sales

    if report.total_impressions > 0:
        report.total_ctr = round(report.total_clicks / report.total_impressions * 100, 2)
    if report.total_cost > 0 and report.total_conversion_sales > 0:
        report.total_roas = round(report.total_conversion_sales / report.total_cost * 100, 1)


def generate_coupang_insights(report: CoupangReport) -> list[str]:
    """쿠팡 광고 인사이트 생성 (종합 분석)."""
    insights = []

    if not report.total_cost and not report.campaigns:
        insights.append("[정보] 분석할 데이터가 충분하지 않습니다.")
        return insights

    # ── 1. 핵심 성과 요약 ──
    if report.total_cost > 0:
        cpc = report.total_cost / max(report.total_clicks, 1)
        cvr = report.total_conversions / max(report.total_clicks, 1) * 100
        insights.append(
            f"[핵심 요약] 총 광고비 {report.total_cost:,.0f}원 집행, "
            f"노출 {report.total_impressions:,}회 / 클릭 {report.total_clicks:,}회 / 전환 {report.total_conversions:,}건 달성"
        )

    # ── 2. 효율 지표 ──
    parts = []
    if report.total_clicks > 0:
        cpc_val = round(report.total_cost / report.total_clicks)
        parts.append(f"CPC {cpc_val:,.0f}원")
    if report.total_impressions > 0:
        ctr_val = round(report.total_clicks / report.total_impressions * 100, 2)
        parts.append(f"CTR {ctr_val:.2f}%")
    if report.total_clicks > 0 and report.total_conversions > 0:
        cvr_val = round(report.total_conversions / report.total_clicks * 100, 1)
        parts.append(f"CVR {cvr_val:.1f}%")
    if report.total_conversions > 0:
        cpa_val = round(report.total_cost / report.total_conversions)
        parts.append(f"CPA {cpa_val:,.0f}원")
    if parts:
        insights.append(f"[효율 지표] {' / '.join(parts)}")

    # ── 3. ROAS 분석 ──
    if report.total_roas > 0:
        if report.total_roas >= 500:
            grade = "매우 우수"
            comment = "광고 효율이 탁월합니다. 예산 확대를 적극 검토하세요."
        elif report.total_roas >= 300:
            grade = "우수"
            comment = "안정적인 수익 구조입니다. 현 전략을 유지하면서 점진적 확대를 고려하세요."
        elif report.total_roas >= 150:
            grade = "양호"
            comment = "수익이 발생하고 있으나, 비효율 캠페인 정리로 더 개선할 수 있습니다."
        elif report.total_roas >= 100:
            grade = "보통 (손익분기)"
            comment = "손익분기 수준입니다. 비효율 키워드/상품 정리가 필요합니다."
        else:
            grade = "주의 필요 (적자)"
            comment = "광고비 대비 매출이 부족합니다. 전환율 개선 또는 입찰가 하향 조정이 시급합니다."
        insights.append(
            f"[ROAS 분석] ROAS {report.total_roas:.0f}% ({grade}) - "
            f"광고비 {report.total_cost:,.0f}원 → 전환매출 {report.total_conversion_sales:,.0f}원. {comment}"
        )

    # ── 4. 전환 분석 ──
    if report.total_conversions > 0 and report.total_clicks > 0:
        cvr = round(report.total_conversions / report.total_clicks * 100, 1)
        cpa = round(report.total_cost / report.total_conversions)
        if cvr >= 5:
            cvr_comment = "높은 전환율로 상품 경쟁력이 좋습니다."
        elif cvr >= 2:
            cvr_comment = "양호한 전환율입니다."
        elif cvr >= 1:
            cvr_comment = "평균 수준의 전환율이며, 상세페이지 개선으로 향상 가능합니다."
        else:
            cvr_comment = "낮은 전환율입니다. 상품 상세페이지, 리뷰, 가격 경쟁력을 점검하세요."
        insights.append(
            f"[전환 분석] 전환율 {cvr:.1f}%, CPA(전환당 비용) {cpa:,.0f}원. {cvr_comment}"
        )

    # ── 5. 매출 분석 ──
    if report.total_sales > 0 and report.total_conversion_sales > 0:
        ad_share = round(report.total_conversion_sales / report.total_sales * 100, 1)
        insights.append(
            f"[매출 분석] 전체매출 {report.total_sales:,.0f}원 중 광고 전환매출 {report.total_conversion_sales:,.0f}원 "
            f"(광고 기여율 {ad_share:.1f}%)"
        )

    # ── 6. 캠페인별 분석 ──
    if report.campaigns:
        sorted_camps = sorted(report.campaigns, key=lambda c: c.cost, reverse=True)
        top = sorted_camps[0]
        cost_share = round(top.cost / max(report.total_cost, 1) * 100, 1)
        insights.append(
            f"[캠페인 분석] 최고 비용 캠페인: '{top.campaign_name}' "
            f"(광고비 {top.cost:,.0f}원, 전체의 {cost_share:.0f}%, ROAS {top.roas:.0f}%)"
        )

        inefficient = [c for c in report.campaigns if c.cost > 0 and 0 < c.roas < 100]
        if inefficient:
            names = ", ".join(f"'{c.campaign_name}'({c.roas:.0f}%)" for c in inefficient[:3])
            total_waste = sum(c.cost for c in inefficient)
            insights.append(
                f"[비효율 캠페인] ROAS 100% 미만: {names} → 총 {total_waste:,.0f}원 적자. "
                f"입찰가 조정 또는 키워드 재검토 필요"
            )

    # ── 7. 상품별 분석 ──
    if report.products:
        sorted_prods = sorted(report.products, key=lambda p: p.conversion_sales, reverse=True)
        if sorted_prods and sorted_prods[0].conversion_sales > 0:
            top = sorted_prods[0]
            insights.append(
                f"[최고 매출 상품] '{top.product_name[:25]}' - "
                f"전환매출 {top.conversion_sales:,.0f}원, ROAS {top.roas:.0f}%"
            )

        no_conv = [p for p in report.products if p.clicks >= 10 and p.conversions == 0]
        if no_conv:
            total_wasted = sum(p.cost for p in no_conv)
            insights.append(
                f"[낭비 상품] 클릭 10회 이상 전환 0인 상품 {len(no_conv)}개 "
                f"(낭비 광고비 {total_wasted:,.0f}원) → 상세페이지/가격 점검 필요"
            )

    # ── 8. 키워드 분석 ──
    if report.keywords:
        waste_kw = [k for k in report.keywords if k.clicks >= 5 and k.conversions == 0 and k.cost > 0]
        if waste_kw:
            waste_total = sum(k.cost for k in waste_kw)
            examples = ", ".join(f"'{k.keyword}'" for k in waste_kw[:5])
            insights.append(
                f"[낭비 키워드] 클릭 발생 전환 0인 키워드 {len(waste_kw)}개 "
                f"(총 {waste_total:,.0f}원): {examples} → 제외 키워드 등록 검토"
            )

    # ── 9. 추천 액션 ──
    actions = []
    if report.total_roas > 300:
        actions.append("ROAS 우수 캠페인 예산 확대 검토")
    if report.total_clicks > 0 and report.total_conversions / max(report.total_clicks, 1) * 100 < 1:
        actions.append("전환율 개선을 위한 상세페이지 최적화")
    inefficient_camps = [c for c in report.campaigns if c.cost > 0 and 0 < c.roas < 100]
    if inefficient_camps:
        actions.append(f"ROAS 100% 미만 캠페인 {len(inefficient_camps)}개 입찰 전략 재검토")
    if report.total_ctr < 0.1 and report.total_impressions > 10000:
        actions.append("CTR 개선을 위한 광고 이미지/제목 최적화")
    if actions:
        insights.append(f"[추천 액션] {' / '.join(actions[:4])}")

    return insights


# ── DOM / API 데이터에서 리포트 생성 ──

def parse_dom_data(dom_data: dict, date_from: date = None, date_to: date = None) -> CoupangReport:
    """
    대시보드 DOM에서 스크래핑한 데이터로 CoupangReport 생성.
    """
    report = CoupangReport(
        report_date=date_from or date.today(),
        date_from=date_from,
        date_to=date_to,
    )

    # 테이블에서 실제 조회 기간 감지
    import re as _re
    tables = dom_data.get("_tables", [])
    for table in tables:
        if table and len(table) > 1:
            # 첫 데이터 행에서 기간 형식 "YYYY/MM/DD~YYYY/MM/DD" 탐색
            for row in table[1:]:
                if row:
                    cell0 = str(row[0]) if row else ""
                    dm = _re.match(r'(\d{4}/\d{2}/\d{2})~(\d{4}/\d{2}/\d{2})', cell0)
                    if dm:
                        try:
                            from datetime import datetime as _dt
                            first_from = _dt.strptime(dm.group(1), "%Y/%m/%d").date()
                            if not report.date_from or first_from < report.date_from:
                                report.date_from = first_from
                        except Exception:
                            pass
            # 마지막 데이터 행에서 종료 날짜
            for row in reversed(table[1:]):
                cell0 = str(row[0]) if row else ""
                if cell0 in ("전체", "합계"):
                    continue
                dm2 = _re.match(r'(\d{4}/\d{2}/\d{2})~(\d{4}/\d{2}/\d{2})', cell0)
                if dm2:
                    try:
                        last_to = _dt.strptime(dm2.group(2), "%Y/%m/%d").date()
                        if not report.date_to or last_to > report.date_to:
                            report.date_to = last_to
                    except Exception:
                        pass
                    break
    if report.date_from and report.date_to:
        logger.info(f"쿠팡 실제 조회 기간: {report.date_from} ~ {report.date_to}")

    # KPI 데이터 매핑
    kpi_map = {
        "ad_cost": "total_cost",
        "total_ad_cost": "total_cost",
        "conv_revenue": "total_conversion_sales",
        "total_revenue": "total_sales",
        "conversions": "total_conversions",
        "orders": "total_conversions",  # 주문수도 전환수로 사용
        "impressions": "total_impressions",
        "clicks": "total_clicks",
    }

    for dom_key, report_key in kpi_map.items():
        val = dom_data.get(dom_key)
        if val:
            current = getattr(report, report_key, 0)
            new_val = _safe_int(val)
            # 더 큰 값 사용 (이미 설정된 값이 있으면 비교)
            if new_val > current:
                setattr(report, report_key, new_val)

    # CTR, ROAS, 전환율 등 비율 값
    for key, attr in [("ctr", "total_ctr"), ("roas", "total_roas")]:
        val = dom_data.get(key)
        if val:
            setattr(report, attr, _safe_float(val))

    # 테이블 데이터에서 KPI 및 캠페인 정보 추출
    tables = dom_data.get("_tables", [])
    for table in tables:
        if not table or len(table) < 2:
            continue
        headers = table[0]

        # 헤더에서 "캠페인" 관련 컬럼 확인
        header_text = " ".join(str(h) for h in headers if h)
        has_campaign_col = any("캠페인" in str(h) and "개" in str(h) for h in headers if h)
        has_period_col = any("기간" in str(h) for h in headers if h)
        has_metrics = any("노출수" in str(h) for h in headers if h)

        if has_campaign_col and has_metrics:
            # 캠페인별 성과 테이블: 합계행에서 KPI 추출
            _parse_performance_table(report, headers, table[1:])

        elif has_period_col and has_metrics:
            # 기간별 성과 테이블: 합계행에서 KPI 추출
            _parse_performance_table(report, headers, table[1:])

        elif any("캠페인" in str(h) for h in headers if h):
            _parse_table_to_campaigns(report, headers, table[1:])

    # 캠페인명 매칭: _campaign_names 리스트를 캠페인에 할당
    campaign_names = dom_data.get("_campaign_names", [])
    if campaign_names and report.campaigns:
        # 캠페인명에서 헤더 텍스트 등 필터링
        filtered_names = [n for n in campaign_names
                          if n and not any(kw in n for kw in ["캠페인", "합계", "전체", "노출수"])]
        logger.info(f"캠페인명 {len(filtered_names)}개: {filtered_names[:5]}")
        for i, camp in enumerate(report.campaigns):
            if i < len(filtered_names) and not camp.campaign_name:
                camp.campaign_name = filtered_names[i]

    # ROAS/CTR 보정 계산
    if report.total_ctr == 0 and report.total_impressions > 0 and report.total_clicks > 0:
        report.total_ctr = round(report.total_clicks / report.total_impressions * 100, 2)
    if report.total_roas == 0 and report.total_cost > 0 and report.total_conversion_sales > 0:
        report.total_roas = round(report.total_conversion_sales / report.total_cost * 100, 1)

    report.insights = generate_coupang_insights(report)

    logger.info(
        f"DOM 파싱 완료: 비용={report.total_cost:,}원, "
        f"클릭={report.total_clicks:,}, 전환={report.total_conversions:,}, "
        f"ROAS={report.total_roas:.0f}%"
    )
    return report


def parse_api_data(api_data: list, date_from: date = None, date_to: date = None) -> CoupangReport:
    """
    캡처한 API 응답 데이터에서 CoupangReport 생성.
    """
    report = CoupangReport(
        report_date=date_from or date.today(),
        date_from=date_from,
        date_to=date_to,
    )

    for item in api_data:
        url = item.get("url", "")
        data = item.get("data")
        if not data:
            continue

        # 대시보드 요약 데이터
        if isinstance(data, dict):
            _extract_kpi_from_api(report, data)

        # 캠페인 리스트
        if isinstance(data, list):
            for entry in data:
                if isinstance(entry, dict) and ("campaignName" in entry or "campaign_name" in entry):
                    camp = CoupangCampaignStats(
                        campaign_name=entry.get("campaignName", entry.get("campaign_name", "")),
                        campaign_type=entry.get("campaignType", entry.get("campaign_type", "")),
                        status=entry.get("status", ""),
                        impressions=_safe_int(entry.get("impressions", entry.get("impCnt", 0))),
                        clicks=_safe_int(entry.get("clicks", entry.get("clkCnt", 0))),
                        cost=_safe_int(entry.get("cost", entry.get("adCost", entry.get("salesAmt", 0)))),
                        conversions=_safe_int(entry.get("conversions", entry.get("cvsCnt", 0))),
                        conversion_sales=_safe_int(entry.get("conversionSales", entry.get("cvsAmt", 0))),
                        roas=_safe_float(entry.get("roas", 0)),
                        total_sales=_safe_int(entry.get("totalSales", entry.get("totalRevenue", 0))),
                    )
                    if camp.cost > 0 and camp.roas == 0 and camp.conversion_sales > 0:
                        camp.roas = round(camp.conversion_sales / camp.cost * 100, 1)
                    if camp.campaign_name:
                        report.campaigns.append(camp)

    _calculate_totals(report)
    report.insights = generate_coupang_insights(report)

    logger.info(
        f"API 파싱 완료: 캠페인={len(report.campaigns)}, "
        f"비용={report.total_cost:,}원, ROAS={report.total_roas:.0f}%"
    )
    return report


def _extract_kpi_from_api(report: CoupangReport, data: dict):
    """API 응답에서 KPI 값을 추출하여 report에 반영."""
    kpi_keys = {
        "totalCost": "total_cost",
        "adCost": "total_cost",
        "totalImpressions": "total_impressions",
        "impCnt": "total_impressions",
        "totalClicks": "total_clicks",
        "clkCnt": "total_clicks",
        "totalConversions": "total_conversions",
        "cvsCnt": "total_conversions",
        "conversionSales": "total_conversion_sales",
        "cvsAmt": "total_conversion_sales",
        "totalSales": "total_sales",
        "totalRevenue": "total_sales",
        "roas": "total_roas",
        "ctr": "total_ctr",
    }
    for api_key, attr in kpi_keys.items():
        if api_key in data:
            val = data[api_key]
            if attr in ("total_roas", "total_ctr"):
                setattr(report, attr, _safe_float(val))
            else:
                setattr(report, attr, _safe_int(val))


def _parse_table_to_campaigns(report: CoupangReport, headers: list, rows: list):
    """DOM 테이블 데이터에서 캠페인 정보 추출."""
    col_idx = {}
    for i, h in enumerate(headers):
        h_str = str(h).strip()
        if h_str in COLUMN_MAP:
            col_idx[COLUMN_MAP[h_str]] = i

    def _safe_get(row, key):
        idx = col_idx.get(key, -1)
        if idx < 0 or idx >= len(row):
            return None
        return row[idx]

    for row in rows:
        if not row or len(row) < 2:
            continue
        name = _safe_get(row, "campaign_name")
        if not name:
            continue
        name_str = str(name).strip()
        if not name_str or name_str in ("합계", "전체", "총합"):
            continue

        conversions = _safe_int(_safe_get(row, "conversions"))
        orders = _safe_int(_safe_get(row, "orders"))
        if not conversions and orders:
            conversions = orders

        camp = CoupangCampaignStats(
            campaign_name=name_str,
            impressions=_safe_int(_safe_get(row, "impressions")),
            clicks=_safe_int(_safe_get(row, "clicks")),
            cost=_safe_int(_safe_get(row, "cost")),
            conversions=conversions,
            conversion_sales=_safe_int(_safe_get(row, "conversion_sales")),
            roas=_safe_float(_safe_get(row, "roas")),
        )
        if camp.campaign_name:
            report.campaigns.append(camp)


def _parse_performance_table(report: CoupangReport, headers: list, rows: list):
    """
    대시보드 성과 테이블(기간별 + 합계)에서 KPI 추출.
    헤더가 중복될 수 있음 (예: '노출수', '노출수').
    마지막 행이 합계.
    """
    if not rows:
        return

    # 중복 헤더 제거하여 고유 헤더 인덱스 매핑
    # 쿠팡 보고서 헤더는 비교 기간 때문에 모든 헤더가 2번씩 나옴:
    #   [기간, 기간, 노출수, 노출수, 클릭수, 클릭수, ...]
    # 또는 캠페인별 탭일 때:
    #   [캠페인(9)개, 캠페인(9)개, 노출수, 노출수, ..., 기간, 노출수, ...]
    # "기간"은 지표가 아니므로 항상 스킵하고, 나머지는 seen으로 중복 제거
    unique_headers = []
    seen = set()
    for h in headers:
        h_str = str(h).strip()
        if not h_str:
            continue
        # "기간"은 지표가 아니므로 항상 스킵
        if h_str == "기간":
            continue
        # 이미 본 헤더는 비교 기간 중복이므로 스킵
        if h_str in seen:
            continue
        unique_headers.append(h_str)
        seen.add(h_str)

    logger.info(f"성과 테이블 고유 헤더 {len(unique_headers)}개: {unique_headers}")

    # 마지막 행 = 합계 행
    total_row = rows[-1] if rows else []
    if not total_row:
        return

    # "캠페인(N)개" 같은 라벨 헤더는 데이터 열이 아님 → 제외
    import re
    metric_headers = [h for h in unique_headers
                      if not re.match(r"캠페인.*개", h)]
    logger.info(f"성과 테이블 지표 {len(metric_headers)}개, 합계 행 값 {len(total_row)}개")

    # 데이터 행에 캠페인명이 포함된 경우 오프셋 계산
    # metric_headers: [노출수, 클릭수, ...] (10개)
    # total_row: [캠페인명?, 2198496회, 2812회, ...] (10개 또는 11개)
    data_offset = 0
    if len(total_row) > len(metric_headers):
        data_offset = len(total_row) - len(metric_headers)
        logger.info(f"데이터 오프셋 {data_offset} (캠페인명 등 비지표 컬럼 포함)")
    elif len(total_row) < len(metric_headers):
        metric_headers = metric_headers[:len(total_row)]

    # 매핑
    kpi_mapping = {
        "노출수": ("total_impressions", _safe_int),
        "클릭수": ("total_clicks", _safe_int),
        "클릭률": ("total_ctr", _safe_float),
        "클릭당 비용": (None, None),  # CPC (별도 필드 없음)
        "광고 전환 주문수": ("total_conversions", _safe_int),
        "광고 전환 판매수": ("total_conversions", _safe_int),  # 대체 전환 지표
        "전환율": (None, None),
        "집행 광고비": ("total_cost", _safe_int),
        "광고 전환 매출": ("total_conversion_sales", _safe_int),
        "전체 매출": ("total_sales", _safe_int),
        "광고수익률": ("total_roas", _safe_float),
    }

    # 합계 행 파싱
    for i, header in enumerate(metric_headers):
        data_idx = i + data_offset
        if data_idx >= len(total_row):
            break
        val = total_row[data_idx]
        mapping = kpi_mapping.get(header)
        if mapping and mapping[0]:
            attr_name, converter = mapping
            parsed_val = converter(val)
            if parsed_val > 0:
                # 성과 테이블 합계 행은 가장 정확한 소스이므로 항상 덮어쓰기
                setattr(report, attr_name, parsed_val)
                logger.info(f"  {header} -> {attr_name} = {parsed_val}")

    # 캠페인별 데이터 추출 (합계 행 제외)
    camp_mapping = {
        "노출수": ("impressions", _safe_int),
        "클릭수": ("clicks", _safe_int),
        "클릭률": ("ctr", _safe_float),
        "광고 전환 주문수": ("conversions", _safe_int),
        "광고 전환 판매수": ("conversions", _safe_int),
        "집행 광고비": ("cost", _safe_int),
        "광고 전환 매출": ("conversion_sales", _safe_int),
        "광고수익률": ("roas", _safe_float),
    }
    # 첫 행이 합계이고 마지막 행도 합계일 수 있음 → 중간 행들이 개별 캠페인
    if len(rows) > 2:
        # 총계행의 첫 번째 지표값 (비교용 - 첫 행이 총계인지 판별)
        total_first_metric = total_row[data_offset] if data_offset < len(total_row) else ""

        for row in rows[:-1]:  # 마지막(합계) 제외
            if not row or len(row) < 3:
                continue
            # 마지막 총계행과 같은 행 스킵
            if row == rows[-1]:
                continue

            first_val = str(row[0]).strip() if row else ""

            # 날짜 형식 행 스킵 (비교 기간 데이터: "2026/02/01~2026/02/07")
            if re.match(r"\d{4}/", first_val):
                continue

            # "전체" 행 스킵
            if first_val == "전체":
                continue

            # 첫 번째 행이 총계와 같은 지표값이면 스킵 (총괄 행)
            row_first_metric = row[data_offset] if data_offset < len(row) else ""
            if row == rows[0] and rows[0] != rows[1] and row_first_metric == total_first_metric:
                continue

            # 캠페인명 추출 (오프셋이 있으면 첫 번째 값이 캠페인명)
            camp_name = ""
            if data_offset > 0 and len(row) > len(metric_headers):
                camp_name = first_val

            camp = CoupangCampaignStats(campaign_name=camp_name)
            for i, header in enumerate(metric_headers):
                data_idx = i + data_offset
                if data_idx >= len(row):
                    break
                val = row[data_idx]
                cm = camp_mapping.get(header)
                if cm:
                    attr, converter = cm
                    setattr(camp, attr, converter(val))

            # 유효한 캠페인 데이터인지 확인 (비용 또는 클릭이 있어야)
            if camp.cost > 0 or camp.clicks > 0:
                report.campaigns.append(camp)

        if report.campaigns:
            logger.info(f"캠페인 {len(report.campaigns)}개 추출 완료")
